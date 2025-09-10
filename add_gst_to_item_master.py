import re
from typing import Dict, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


def to_pct(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    if s.lower() == "nil":
        return 0.0
    if s.endswith("%"):
        s = s[:-1]
    try:
        return float(s)
    except Exception:
        return None


def digits_only(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"\D", "", str(value))


def normalize_to_upto8(code: Optional[str]) -> str:
    return digits_only(code)[:8]


def build_gst_prefix_maps(gst_xlsx_path: str) -> Dict[int, Dict[str, Tuple[Optional[float], Optional[float]]]]:
    wb = load_workbook(gst_xlsx_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {2: {}, 4: {}, 6: {}, 8: {}}

    header = rows[0]
    try:
        idx_hsn = header.index("HSN Code")
        idx_from = header.index("From Rate")
        idx_to = header.index("To Rate")
    except ValueError as exc:
        raise RuntimeError("GST-Rates.xlsx missing required columns") from exc

    maps: Dict[int, Dict[str, Tuple[Optional[float], Optional[float]]]] = {2: {}, 4: {}, 6: {}, 8: {}}

    for r in rows[1:]:
        if r is None:
            continue
        hsn_cell = r[idx_hsn] if idx_hsn < len(r) else None
        from_rate = to_pct(r[idx_from] if idx_from < len(r) else None)
        to_rate = to_pct(r[idx_to] if idx_to < len(r) else None)
        if hsn_cell is None:
            continue
        for chunk in re.split(r"\s*,\s*", str(hsn_cell)):
            if not chunk:
                continue
            digits = digits_only(chunk)
            if not digits:
                continue
            digits = digits[:8]
            for length in (8, 6, 4, 2):
                if len(digits) >= length:
                    key = digits[:length]
                    prev = maps[length].get(key)
                    if prev is None:
                        maps[length][key] = (from_rate, to_rate)
                    else:
                        pf, pt = prev
                        if pf is None and from_rate is not None:
                            pf = from_rate
                        if pt is None and to_rate is not None:
                            pt = to_rate
                        maps[length][key] = (pf, pt)
    return maps


def enrich_item_master(item_csv_path: str, gst_xlsx_path: str, out_csv_path: str) -> int:
    gst_maps = build_gst_prefix_maps(gst_xlsx_path)

    df = pd.read_csv(item_csv_path)

    colmap = {c.lower(): c for c in df.columns}
    hsn_col = None
    for name in ("hsn", "hsn_code", "hsncode", "hsn code"):
        if name in colmap:
            hsn_col = colmap[name]
            break
    if hsn_col is None:
        for k, v in colmap.items():
            if "hsn" in k:
                hsn_col = v
                break
    if hsn_col is None:
        raise RuntimeError("Could not find an HSN column in Item Master CSV")

    # Vectorized normalization and prefix generation
    digits = df[hsn_col].astype(str).map(digits_only).str.slice(0, 8)
    k8 = digits.str.slice(0, 8)
    k6 = digits.str.slice(0, 6)
    k4 = digits.str.slice(0, 4)
    k2 = digits.str.slice(0, 2)

    # Build separate dicts for old/new by level for fast Series.map
    old8 = {k: v[0] for k, v in gst_maps[8].items()}
    new8 = {k: v[1] for k, v in gst_maps[8].items()}
    old6 = {k: v[0] for k, v in gst_maps[6].items()}
    new6 = {k: v[1] for k, v in gst_maps[6].items()}
    old4 = {k: v[0] for k, v in gst_maps[4].items()}
    new4 = {k: v[1] for k, v in gst_maps[4].items()}
    old2 = {k: v[0] for k, v in gst_maps[2].items()}
    new2 = {k: v[1] for k, v in gst_maps[2].items()}

    # Coalesce from most specific to least
    old_series = (
        k8.map(old8).where(k8.notna())
        .fillna(k6.map(old6))
        .fillna(k4.map(old4))
        .fillna(k2.map(old2))
    )
    new_series = (
        k8.map(new8).where(k8.notna())
        .fillna(k6.map(new6))
        .fillna(k4.map(new4))
        .fillna(k2.map(new2))
    )

    out = df.copy()
    out["old_gst"] = old_series
    out["new_gst"] = new_series

    out.to_csv(out_csv_path, index=False)

    matched = int(out[["old_gst", "new_gst"]].notna().any(axis=1).sum())
    return matched


def enrich_item_master_chunked(item_csv_path: str, gst_xlsx_path: str, out_csv_path: str, chunksize: int = 100_000) -> int:
    """Chunked variant for very large CSVs. Returns matched row count.

    This streams input and output to keep memory usage low.
    """
    gst_maps = build_gst_prefix_maps(gst_xlsx_path)

    # Prepare mapping dicts once
    old8 = {k: v[0] for k, v in gst_maps[8].items()}
    new8 = {k: v[1] for k, v in gst_maps[8].items()}
    old6 = {k: v[0] for k, v in gst_maps[6].items()}
    new6 = {k: v[1] for k, v in gst_maps[6].items()}
    old4 = {k: v[0] for k, v in gst_maps[4].items()}
    new4 = {k: v[1] for k, v in gst_maps[4].items()}
    old2 = {k: v[0] for k, v in gst_maps[2].items()}
    new2 = {k: v[1] for k, v in gst_maps[2].items()}

    total_matched = 0
    header_written = False

    for chunk in pd.read_csv(item_csv_path, chunksize=chunksize):
        colmap = {c.lower(): c for c in chunk.columns}
        hsn_col = None
        for name in ("hsn", "hsn_code", "hsncode", "hsn code"):
            if name in colmap:
                hsn_col = colmap[name]
                break
        if hsn_col is None:
            for k, v in colmap.items():
                if "hsn" in k:
                    hsn_col = v
                    break
        if hsn_col is None:
            raise RuntimeError("Could not find an HSN column in Item Master CSV")

        digits = chunk[hsn_col].astype(str).map(digits_only).str.slice(0, 8)
        k8 = digits.str.slice(0, 8)
        k6 = digits.str.slice(0, 6)
        k4 = digits.str.slice(0, 4)
        k2 = digits.str.slice(0, 2)

        old_series = (
            k8.map(old8).where(k8.notna())
            .fillna(k6.map(old6))
            .fillna(k4.map(old4))
            .fillna(k2.map(old2))
        )
        new_series = (
            k8.map(new8).where(k8.notna())
            .fillna(k6.map(new6))
            .fillna(k4.map(new4))
            .fillna(k2.map(new2))
        )

        out = chunk.copy()
        out["old_gst"] = old_series
        out["new_gst"] = new_series

        total_matched += int(out[["old_gst", "new_gst"]].notna().any(axis=1).sum())

        out.to_csv(out_csv_path, mode="a", index=False, header=not header_written)
        header_written = True

    return total_matched


if __name__ == "__main__":
    import argparse
    import sys
    from pathlib import Path

    def get_default_gst_path() -> Path:
        # If packaged by PyInstaller, read from the temporary bundle dir
        base = Path(getattr(sys, "_MEIPASS", Path(__file__).parent))
        return base / "GST-Rates.xlsx"

    parser = argparse.ArgumentParser(description="Add GST rates to an Item Master CSV using GST-Rates.xlsx")
    parser.add_argument("--input", help="Path to input Item Master CSV")
    parser.add_argument("--gst", help="Path to GST-Rates.xlsx (default: embedded/alongside the app)")
    parser.add_argument("--output", help="Output CSV path (default: <input>_with_gst.csv")
    parser.add_argument("--chunksize", type=int, default=0, help="Process in chunks (e.g. 100000) to reduce memory usage")

    args = parser.parse_args()

    default_gst_path = get_default_gst_path()

    if not args.input:
        try:
            import tkinter as tk
            from tkinter import filedialog, messagebox

            def choose_input():
                path = filedialog.askopenfilename(title="Select Item Master CSV", filetypes=[["CSV files", "*.csv"], ["All files", "*.*"]])
                if path:
                    inp_var.set(path)
                    if not out_var.get():
                        p = Path(path).with_name(Path(path).stem + "_with_gst.csv")
                        out_var.set(str(p))

            def choose_output():
                initial = Path(inp_var.get()).with_name(Path(inp_var.get()).stem + "_with_gst.csv") if inp_var.get() else Path("output_with_gst.csv")
                path = filedialog.asksaveasfilename(title="Save output CSV as", defaultextension=".csv", initialfile=initial.name, filetypes=[["CSV files", "*.csv"]])
                if path:
                    out_var.set(path)

            def run_process():
                inp = inp_var.get().strip()
                outp = out_var.get().strip()
                gstp = gst_var.get().strip()
                if not inp:
                    messagebox.showwarning("Missing", "Please select an input CSV file.")
                    return
                if not outp:
                    messagebox.showwarning("Missing", "Please choose an output CSV path.")
                    return
                try:
                    btn_run.config(state=tk.DISABLED)
                    status_var.set("Processing...")
                    root.update_idletasks()
                    if args.chunksize and args.chunksize > 0:
                        count = enrich_item_master_chunked(item_csv_path=inp, gst_xlsx_path=gstp, out_csv_path=outp, chunksize=args.chunksize)
                    else:
                        count = enrich_item_master(item_csv_path=inp, gst_xlsx_path=gstp, out_csv_path=outp)
                    status_var.set(f"Done. Output saved. Matched rows: {count}")
                    messagebox.showinfo("Done", f"Output saved to:\n{outp}\nMatched rows: {count}")
                except Exception as e:
                    status_var.set("Error")
                    messagebox.showerror("Error", str(e))
                finally:
                    btn_run.config(state=tk.NORMAL)

            root = tk.Tk()
            root.title("GST Rate Enricher")
            root.geometry("560x260")
            root.resizable(False, False)

            inp_var = tk.StringVar()
            out_var = tk.StringVar()
            gst_var = tk.StringVar(value=str(default_gst_path))
            status_var = tk.StringVar(value=f"Using GST file: {Path(gst_var.get()).name}")

            tk.Label(root, text="Item Master CSV:").place(x=20, y=20)
            tk.Entry(root, textvariable=inp_var, width=60).place(x=140, y=20)
            tk.Button(root, text="Browse", command=choose_input).place(x=460, y=16)

            tk.Label(root, text="Output CSV:").place(x=20, y=60)
            tk.Entry(root, textvariable=out_var, width=60).place(x=140, y=60)
            tk.Button(root, text="Save as", command=choose_output).place(x=460, y=56)

            tk.Label(root, text="GST Rates file:").place(x=20, y=100)
            tk.Entry(root, textvariable=gst_var, width=60, state="disabled").place(x=140, y=100)

            btn_run = tk.Button(root, text="Run", width=12, command=run_process)
            btn_run.place(x=140, y=150)

            tk.Label(root, textvariable=status_var, fg="gray").place(x=140, y=190)

            root.mainloop()
            raise SystemExit
        except Exception:
            pass

    in_path = Path(args.input)
    out_path = Path(args.output) if args.output else in_path.with_name(in_path.stem + "_with_gst.csv")
    gst_path = Path(args.gst) if args.gst else default_gst_path

    if args.chunksize and args.chunksize > 0:
        matched = enrich_item_master_chunked(
            item_csv_path=str(in_path),
            gst_xlsx_path=str(gst_path),
            out_csv_path=str(out_path),
            chunksize=args.chunksize,
        )
    else:
        matched = enrich_item_master(
            item_csv_path=str(in_path),
            gst_xlsx_path=str(gst_path),
            out_csv_path=str(out_path),
        )
    print(f"Matched rows: {matched}")
